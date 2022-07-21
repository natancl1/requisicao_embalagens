'''
Esse código lança as requisições na base de dados histórico.
'''

import os
import sqlite3

from datetime import datetime
from openpyxl import load_workbook



def conectar_db():
    # Criação/conexão do banco de dados e da tabela "requisicoes"
    con = sqlite3.connect('database.db')
    cur = con.cursor()
    cur.execute('''CREATE TABLE IF NOT EXISTS requisicoes 
                    (requisicao integer, data text, cod_embalagem text, embalagem text, 
                    qtd_solicitada real, hora text, qtd_programada real, qtd_realizada real, 
                    saldo real, percent_consumo real, requisitante text, lote text)''')
    return con, cur


# Função para obter as informações que aparecem nos dados de info quando clica no botão com 'i' na janela principal
def get_info(cod_embalagem, qtd_solicitada, data):
    con, cur = conectar_db()

    embalagem = list(cur.execute('SELECT embalagem FROM ficha_tecnica WHERE cod_embalagem = "' + cod_embalagem + '"'))
    if embalagem == []:
        embalagem = 'Em branco'
    else:
        embalagem = embalagem[0][0]

    qtd_programada = list(cur.execute('''SELECT SUM(ficha_tecnica.consumo_por_caixa * programacao.cxs)
                                    FROM programacao
                                    LEFT JOIN ficha_tecnica ON ficha_tecnica.cod_produto = programacao.cod
                                    WHERE ficha_tecnica.cod_embalagem = "''' + cod_embalagem + '''"
                                    GROUP BY ficha_tecnica.cod_embalagem'''))
    if qtd_programada == []:
        qtd_programada = 0
    else:
        qtd_programada = qtd_programada[0][0]

    qtd_realizada = list(cur.execute('SELECT SUM(qtd_solicitada) FROM requisicoes WHERE cod_embalagem = "' + 
                                cod_embalagem + '" AND data = "' + data + '"'))
    if qtd_realizada == [(None,)] or qtd_realizada == []:
        qtd_realizada = qtd_solicitada
    else:
        qtd_realizada = qtd_realizada[0][0] + qtd_solicitada

    saldo = qtd_realizada - qtd_programada

    status = 'LIBERADO'
    if saldo > 0:
        status = 'BLOQUEADO'

    try:
        percent_consumo = qtd_realizada / qtd_programada
    except ZeroDivisionError:
        percent_consumo = qtd_realizada
    
    con.close()
    return (embalagem, status, qtd_programada, qtd_realizada, saldo, percent_consumo)


# Funções para salvar as requisições no banco de dados e imprimir a etiqueta
def salvar_imprimir(cod_embalagem, qtd_solicitada, lote, requisitante, con, cur):

    data = datetime.today().strftime('%d/%m/%Y')
    embalagem, status, qtd_programada, qtd_realizada, saldo, percent_consumo = get_info(cod_embalagem, qtd_solicitada, data)

    requisicao = list(cur.execute('SELECT requisicao FROM requisicoes ORDER BY requisicao DESC LIMIT 1'))
    if requisicao == []:
        requisicao = 1
    else:
        requisicao = requisicao[0][0] + 1

    hora = datetime.today().strftime('%H:%M')

    cur.execute('INSERT INTO requisicoes VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)', 
                (requisicao, data, cod_embalagem, embalagem, qtd_solicitada, hora, qtd_programada, 
                qtd_realizada, saldo, percent_consumo, requisitante, lote))
    con.commit()
    con.close()

    wb = load_workbook('etiqueta.xlsm', keep_vba=True)
    ws = wb['etiqueta']
    ws['D1'] = requisicao
    ws['C3'] = data
    ws['C9'] = cod_embalagem
    ws['C4'] = embalagem
    ws['C5'] = qtd_solicitada
    ws['C7'] = datetime.today().strftime('%H:%M:%S')
    ws['D3'] = qtd_programada
    ws['D5'] = qtd_realizada
    ws['D7'] = saldo
    ws['C6'] = requisitante
    ws['B2'] = lote
    wb.save('etiqueta.xlsm')
    wb.close()

    os.startfile("etiqueta.xlsm", "print")

def fim_requisicao(cod_embalagem, qtd_solicitada, lote, requisitante):
    data = datetime.today().strftime('%d/%m/%Y')
    _, status, _, _, _, _ = get_info(cod_embalagem, qtd_solicitada, data)
    con, cur = conectar_db()
    if status == 'BLOQUEADO':
        import tkinter as tk
        from tkinter import messagebox
        
        root = tk.Tk()
        root.withdraw()
        root.iconbitmap('Logo-Uniaves_2020.ico')

        resposta = messagebox.askquestion("Atenção!", "Saldo indisponível, deseja continuar?", icon='warning')

        if resposta == 'yes':
            salvar_imprimir(cod_embalagem, qtd_solicitada, lote, requisitante, con, cur)
        else:
            con.close()
    else:
        salvar_imprimir(cod_embalagem, qtd_solicitada, lote, requisitante, con, cur)
