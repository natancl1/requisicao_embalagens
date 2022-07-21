'''
Rotina para atualizar a tabela de ficha técnica dos produtos para calcular o consumo programado de embalagens.
'''

import sqlite3

from tkinter import Tk, filedialog
from openpyxl import load_workbook

# Criação/conexão do banco de dados e da tabela "ficha_tecnica"
con = sqlite3.connect('database.db')
cur = con.cursor()
cur.execute('DROP TABLE IF EXISTS ficha_tecnica')
cur.execute('''CREATE TABLE IF NOT EXISTS ficha_tecnica 
                (cod_produto integer, produto text, kg_por_caixa real, 
                cod_embalagem text, embalagem text, consumo_por_kg real, consumo_por_caixa real)''')

# Rotina para abrir a planilha ler as informações, inserir e salvar no banco de dados
root = Tk()
root.withdraw()
root.iconbitmap('Logo-Uniaves_2020.ico')
planilha = filedialog.askopenfilename(filetypes=[("Excel file", "*.xlsx")])
wb = load_workbook(planilha, data_only=True) # Workbook - Pasta de trabalho
ws = wb['ficha_tecnica'] # Worksheet - Planilha
cod_produto = [i.value for i in ws['A']]
produto = [i.value for i in ws['B']]
kg_por_caixa = [i.value for i in ws['C']]
cod_embalagem = [i.value for i in ws['D']]
embalagem = [i.value for i in ws['E']]
consumo_por_kg = [i.value for i in ws['F']]
consumo_por_caixa = []
for i in range(len(cod_produto)):
    try:
        consumo_por_caixa.append(kg_por_caixa[i] * consumo_por_kg[i])
    except TypeError:
        consumo_por_caixa.append(0)

ficha = list(zip(cod_produto, produto, kg_por_caixa, cod_embalagem, embalagem, consumo_por_kg, consumo_por_caixa))[1:]
cur.executemany('INSERT INTO ficha_tecnica VALUES (?, ?, ?, ?, ?, ?, ?)', ficha)
con.commit()
con.close()
