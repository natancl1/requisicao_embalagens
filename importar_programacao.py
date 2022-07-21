'''
Rotina para atualizar a tabela de programação da produção que servirá para calcular o consumo de embalagem programado.
'''

import sqlite3

from tkinter import Tk, filedialog
from openpyxl import load_workbook

# Criação/conexão do banco de dados e da tabela "programacao"
con = sqlite3.connect('database.db')
cur = con.cursor()
cur.execute('DROP TABLE IF EXISTS programacao')
cur.execute('CREATE TABLE IF NOT EXISTS programacao (cod integer, produto text, cxs integer)')

# Rotina para abrir a planilha ler as informações, inserir e salvar no banco de dados
root = Tk()
root.withdraw()
root.iconbitmap('Logo-Uniaves_2020.ico')
planilha = filedialog.askopenfilename(filetypes=[("Excel file", "*.xlsx")])
wb = load_workbook(planilha, data_only=True) # Workbook - Pasta de trabalho
ws = wb[wb.sheetnames[0]] # Worksheet - Planilha
cod = [i.value for i in ws['A']]     #
produto = [i.value for i in ws['B']] # Colunas
cxs = [i.value for i in ws['D']]     #
mix_ideal = list(zip(cod, produto, cxs))[2:-1]
cur.executemany('INSERT INTO programacao VALUES (?, ?, ?)', mix_ideal)
con.commit()
con.close()
